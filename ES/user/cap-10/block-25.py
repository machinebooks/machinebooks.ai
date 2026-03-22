# Extraído de: LibroUsuario/cap-10-construir-tu-conector-mcp.md
@servidor.tool()
async def consultar_hoja(hoja: str, filtro: str) -> str:
    """
    Consulta datos en la hoja de cálculo de seguimiento de proyectos.
    Parámetros:
    - hoja: nombre de la pestaña (Proyectos, Clientes, Presupuestos)
    - filtro: texto para filtrar resultados (busca en todas las columnas)
    """
    import openpyxl

    libro = openpyxl.load_workbook("seguimiento-proyectos.xlsx",
                                    read_only=True)
    if hoja not in libro.sheetnames:
        return f"La hoja '{hoja}' no existe. Hojas disponibles: {libro.sheetnames}"

    ws = libro[hoja]
    filas = list(ws.iter_rows(values_only=True))
    cabeceras = filas[0]
    datos = filas[1:]

    # Filtrar
    if filtro:
        filtro_lower = filtro.lower()
        datos = [f for f in datos
                 if any(filtro_lower in str(v).lower() for v in f if v)]

    resultado = f"Hoja: {hoja} | Filtro: '{filtro}'\n"
    resultado += " | ".join(str(c) for c in cabeceras) + "\n"
    resultado += "-" * 60 + "\n"
    for fila in datos[:50]:
        resultado += " | ".join(str(v) if v else "" for v in fila) + "\n"

    return resultado
