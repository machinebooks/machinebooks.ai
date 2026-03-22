# Extraído de: LibroCISO/cap-19-dashboards-copiloto.md
# Ejemplo didáctico: endpoint de exportación multi-formato
from fastapi.responses import Response
from openpyxl import Workbook
from weasyprint import HTML
from docx import Document
from io import BytesIO

@router.get("/treatments/export")
async def export_treatments(
    format: str = Query(..., regex="^(xlsx|pdf|docx)$"),
    legal_basis: str | None = Query(None),
    dpia_status: str | None = Query(None),
    db: AsyncSession = Depends(get_db),
    corporate_id: int = Depends(get_tenant),
):
    """
    Exporta la lista de tratamientos en el formato solicitado.
    Aplica los mismos filtros que la tabla del frontend.
    """
    # Recuperar tratamientos con los filtros aplicados
    treatments = await _get_filtered_treatments(
        db, corporate_id, legal_basis, dpia_status
    )

    # Si hay más de 500 registros, delegar a Celery
    if len(treatments) > 500:
        task = export_treatments_task.delay(
            corporate_id=corporate_id,
            format=format,
            filters={"legal_basis": legal_basis, "dpia_status": dpia_status},
        )
        return {"task_id": task.id, "status": "processing"}

    # Exportación directa para conjuntos pequeños
    if format == "xlsx":
        return _export_xlsx(treatments)
    elif format == "pdf":
        return _export_pdf(treatments)
    elif format == "docx":
        return _export_docx(treatments)


def _export_xlsx(treatments: list) -> Response:
    """Genera un Excel con openpyxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Actividades de Tratamiento"

    # Cabeceras según Art. 30 RGPD
    headers = [
        "Nombre", "Base jurídica", "Finalidad", "Categorías de datos",
        "Categorías de interesados", "Destinatarios", "Transferencias int.",
        "Plazos de conservación", "Estado DPIA", "Última revisión",
    ]
    ws.append(headers)

    # Estilo de cabecera
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E3A5F", fill_type="solid")

    # Datos
    for t in treatments:
        ws.append([
            t.name, t.legal_basis, t.purpose,
            ", ".join(t.data_categories),
            ", ".join(t.subject_categories),
            ", ".join(t.recipients) if t.recipients else "—",
            "Sí" if t.international_transfers else "No",
            t.retention_period or "No definido",
            t.dpia_status, t.last_review.isoformat(),
        ])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=RAT_export.xlsx"},
    )


def _export_pdf(treatments: list) -> Response:
    """Genera un PDF con WeasyPrint desde una plantilla HTML."""
    # Renderizar plantilla Jinja2 con los datos
    html_content = render_template(
        "exports/treatments_report.html",
        treatments=treatments,
        generated_at=datetime.now().isoformat(),
        title="Registro de Actividades de Tratamiento (Art. 30 RGPD)",
    )
    pdf_bytes = HTML(string=html_content).write_pdf()

    return Response(
        content=pdf_bytes,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=RAT_report.pdf"},
    )
